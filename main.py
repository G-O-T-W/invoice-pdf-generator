import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path

from openpyxl.styles.builtins import total

# Define the path to the directory containing the CSV files
filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    # Create a new PDF document
    pdf = FPDF(orientation="p", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False, margin=0)
    pdf.add_page()

    # Set invoice header
    filename = Path(filepath).stem
    invoice_nr, date = filename.split("-")
    pdf.set_font("Times", size=18, style="B")
    pdf.cell(w=0, h=9, txt=f"Invoice nr.{invoice_nr}", align="l", ln=1)
    pdf.cell(w=0, h=9, txt=f"Date: {date}", align="l", ln=1)
    pdf.ln(9)

    # Set table header
    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]

    pdf.set_font("Times", size=8, style="B")

    pdf.cell(w=30, h=10, txt=columns[0], border=1)
    pdf.cell(w=70, h=10, txt=columns[1], border=1)
    pdf.cell(w=30, h=10, txt=columns[2], border=1)
    pdf.cell(w=30, h=10, txt=columns[3], border=1)
    pdf.cell(w=30, h=10, txt=columns[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font("Times", size=8)
        pdf.set_text_color(60, 60, 60)

        pdf.cell(w=30, h=10, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=10, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=10, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=10, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=10, txt=str(row["total_price"]), border=1, ln=1)

    # Set total price at the bottom of the PDF
    total = df["total_price"].sum()

    pdf.set_font("Times", size=8, style="B")
    pdf.cell(w=30, h=10, txt="", border=1)
    pdf.cell(w=70, h=10, txt="", border=1)
    pdf.cell(w=30, h=10, txt="", border=1)
    pdf.cell(w=30, h=10, txt="", border=1)
    pdf.cell(w=30, h=10, txt=str(total), border=1, ln=1)

    pdf.set_text_color(0, 0, 0)

    # Set up summary of invoice
    pdf.set_font("Times", size=12, style="B")
    pdf.cell(w=0, h=10, txt=f"The total price is {total}.", ln=1)
    pdf.ln(10)

    # Set up company name and logo
    pdf.set_font("Helvetica", size=14, style="B")
    pdf.cell(w=30, h=10, txt="Created By: ")
    pdf.image("logo.png", w=20)

    pdf.output(f"PDFs/{filename}.pdf")









